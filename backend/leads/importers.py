import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, TextIOWrapper

from django.contrib.auth import get_user_model

from activities.models import Activity
from .assignment import AssignmentService
from .models import Lead, LeadSource

User = get_user_model()


HEADER_ALIASES = {
    "name": {"name", "full name", "client name", "lead name", "customer name"},
    "phone": {"phone", "mobile", "mobile number", "phone number", "contact", "contact number"},
    "email": {"email", "email address", "mail"},
    "city": {"city", "location", "area"},
    "source": {"source", "lead source", "campaign source", "platform"},
    "project_type": {"project type", "project", "service", "requirement", "interest"},
    "budget": {"budget", "estimated budget", "amount", "price"},
    "timeline": {"timeline", "timeframe", "start time", "when"},
    "notes": {"notes", "message", "comment", "description", "remarks"},
}

PROJECT_TYPE_MAP = {
    "residential": Lead.ProjectType.RESIDENTIAL,
    "house": Lead.ProjectType.RESIDENTIAL,
    "home": Lead.ProjectType.RESIDENTIAL,
    "office": Lead.ProjectType.OFFICE,
    "interior": Lead.ProjectType.OFFICE,
    "commercial": Lead.ProjectType.COMMERCIAL,
    "shop": Lead.ProjectType.COMMERCIAL,
    "renovation": Lead.ProjectType.RENOVATION,
    "kitchen": Lead.ProjectType.RENOVATION,
    "villa": Lead.ProjectType.VILLA,
}

TIMELINE_MAP = {
    "immediate": Lead.Timeline.IMMEDIATE,
    "now": Lead.Timeline.IMMEDIATE,
    "urgent": Lead.Timeline.IMMEDIATE,
    "1 month": Lead.Timeline.ONE_MONTH,
    "one month": Lead.Timeline.ONE_MONTH,
    "3 months": Lead.Timeline.THREE_MONTHS,
    "three months": Lead.Timeline.THREE_MONTHS,
    "not decided": Lead.Timeline.NOT_DECIDED,
}


class LeadImportService:
    @staticmethod
    def import_file(*, uploaded_file, default_source=None, assigned_to_id=None, auto_assign=True, created_by=None):
        rows = LeadImportService._read_rows(uploaded_file)
        assignee = User.objects.filter(id=assigned_to_id, role=User.Role.EXECUTIVE).first() if assigned_to_id else None
        summary = {"created": 0, "skipped": 0, "errors": []}

        for row_number, raw in enumerate(rows, start=2):
            row = LeadImportService._normalize_row(raw)
            try:
                name = (row.get("name") or "").strip()
                phone = LeadImportService._clean_phone(row.get("phone"))
                if not name or not phone:
                    raise ValueError("Name and phone are required.")
                if Lead.objects.filter(phone=phone).exists():
                    summary["skipped"] += 1
                    summary["errors"].append({"row": row_number, "reason": f"Duplicate phone skipped: {phone}"})
                    continue

                source = LeadImportService._resolve_source(row.get("source"), default_source)
                lead = Lead.objects.create(
                    name=name,
                    phone=phone,
                    email=(row.get("email") or "").strip(),
                    city=(row.get("city") or "").strip(),
                    source=source,
                    project_type=LeadImportService._map_project_type(row.get("project_type")),
                    budget=LeadImportService._parse_budget(row.get("budget")),
                    timeline=LeadImportService._map_timeline(row.get("timeline")),
                    notes=(row.get("notes") or "").strip(),
                    assigned_to=assignee,
                    assignment_method="manual_import" if assignee else "",
                    assignment_reason=f"Assigned during import to {assignee.email}." if assignee else "",
                    created_by=created_by,
                )
                if auto_assign and not assignee:
                    AssignmentService.assign_if_needed(lead, user=created_by)
                Activity.objects.create(
                    lead=lead,
                    type=Activity.Type.NOTE,
                    description="Lead imported from spreadsheet.",
                    metadata={"filename": uploaded_file.name, "row": row_number},
                    created_by=created_by,
                )
                summary["created"] += 1
            except Exception as exc:
                summary["skipped"] += 1
                summary["errors"].append({"row": row_number, "reason": str(exc)})
        return summary

    @staticmethod
    def _read_rows(uploaded_file):
        name = uploaded_file.name.lower()
        if name.endswith(".csv"):
            wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
            return list(csv.DictReader(wrapper))
        if name.endswith(".xlsx"):
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            return [dict(zip(headers, values)) for values in rows[1:] if any(values)]
        raise ValueError("Upload a .xlsx or .csv file.")

    @staticmethod
    def _normalize_row(raw):
        normalized = {}
        for header, value in raw.items():
            clean_header = str(header or "").strip().lower()
            for target, aliases in HEADER_ALIASES.items():
                if clean_header in aliases:
                    normalized[target] = "" if value is None else str(value).strip()
                    break
        return normalized

    @staticmethod
    def _resolve_source(value, default_source):
        if value:
            source, _ = LeadSource.objects.get_or_create(name=value.strip())
            return source
        if default_source:
            return default_source
        source, _ = LeadSource.objects.get_or_create(name="Spreadsheet Import")
        return source

    @staticmethod
    def _clean_phone(value):
        phone = str(value or "").strip()
        if phone.endswith(".0"):
            phone = phone[:-2]
        return phone

    @staticmethod
    def _parse_budget(value):
        text = str(value or "").replace(",", "").replace("Rs.", "").replace("₹", "").strip()
        if not text:
            return None
        try:
            amount = Decimal(text)
        except InvalidOperation:
            return None
        return amount if amount > 0 else None

    @staticmethod
    def _map_project_type(value):
        text = str(value or "").lower()
        for key, project_type in PROJECT_TYPE_MAP.items():
            if key in text:
                return project_type
        return Lead.ProjectType.RENOVATION

    @staticmethod
    def _map_timeline(value):
        text = str(value or "").lower()
        for key, timeline in TIMELINE_MAP.items():
            if key in text:
                return timeline
        return Lead.Timeline.NOT_DECIDED
